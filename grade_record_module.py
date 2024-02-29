import openpyxl
from openpyxl.styles import Font
import datetime

class ExcelManager:
    def __init__(self):
        self.wb = None # workbook
        self.ws = None # worksheet
        self.is_excel_open = False
        self.subject_name = '' # subject name
        self.save_path = ''

    def create_excel(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active #指定当前显示（活动）的sheet对象
        self.ws.title = "Score Sheet"

        # 添加表头
        self.ws['A1'] = "Name"
        self.ws['B1'] = "Start Time"
        self.ws['C1'] = "Submission Time"
        self.ws['D1'] = "Score"

        # 设置表头字体加粗
        for cell in self.ws["1:1"]:
            cell.font = Font(bold=True)

    def write_to_excel(self, name, start_time, submit_time, score):
        if not self.is_excel_open:
            print("Excel Doesn't Exist Or Open Excel File First !")
            return

        row = self.ws.max_row + 1
        self.ws.cell(row=row, column=1, value=name)
        self.ws.cell(row=row, column=2, value=start_time)
        self.ws.cell(row=row, column=3, value=submit_time)
        self.ws.cell(row=row, column=4, value=score)

    def save_and_close_excel(self):
        if not self.is_excel_open:
            print("Excel Doesn't Exist Or Open Excel File First !")
            return

        save_name = self.save_path + '/' + self.subject_name + "_Score_Sheet.xlsx"
        self.wb.save(save_name)
        self.wb.close()
        self.is_excel_open = False
        print("Excel File Has Been Saved And Closed at " + save_name)


if __name__ == '__main__':
    # Example
    excel_manager = ExcelManager()
    excel_manager.save_path = 'D:/TestAutoCheck'

    while True:
        signal = int(input("Please enter the signal number \n(1: generate Excel table; 2: write one record; 3: save and close Excel; other numbers: exit not saved !): "))

        if signal == 1:
            subject_name = input("Enter Subject Name: ")
            excel_manager.subject_name = subject_name
            excel_manager.create_excel()
            excel_manager.is_excel_open = True
            print("Excel File Has Been Generated")

        elif signal == 2:
            if not excel_manager.is_excel_open:
                print("Please Generate Excel File First !")
                continue

            name = input("Enter Name: ")
            start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") # current time
            submit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") # current time
            score = int(input("Enter Score: "))

            excel_manager.write_to_excel(name, start_time, submit_time, score)
            print("The Content Has Been Written Into The Excel Table")

        elif signal == 3:
            excel_manager.save_and_close_excel()
            break

        else:
            print('----------------------Exit------------------------')
            break


